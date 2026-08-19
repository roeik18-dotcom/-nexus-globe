"""Master-config retrieval backend (Human + Music).

Small, query-scoped reader for the REAL master workbooks (the canonical xlsx in
Dropbox), used by domain_router's HUMAN_CONFIG / MUSIC_CONFIG routes. It:
  - selects the canonical (highest-version, schema-valid) Human master
  - reads only the rows relevant to the query (never the whole workbook into the prompt)
  - preserves provenance (Canonical_ID / Atomic_ID / Source_ID / Heading)
  - returns UNKNOWN when the master file / required columns are unavailable
  - never invents fields, never promotes person.yaml/music.yaml to master

This is NOT a new router or pipeline — just the data reader the existing route calls.
"""
from __future__ import annotations

import os
import re
import unicodedata
from pathlib import Path

_HUMAN_DIR = Path(os.path.expanduser(
    "~/Library/CloudStorage/Dropbox/----text----/+אדם/קונפינג-אדם-מאגר-אב-שלד-היררכי"))
_MUSIC_FILE = Path(os.path.expanduser(
    "~/Library/CloudStorage/Dropbox/----text----/+מוזיקה/MASTER_MUSIC.xlsx"))

_HUMAN_SHEET = "MASTER_UNITS"
_MUSIC_SHEET = "MASTER_MUSIC"
_REQUIRED = ("Canonical_ID", "Atomic_ID", "Original_Text", "Source_ID")

_STOP = set("מה של את זה עם גם או אבל כי אם על אל לא כן יש אין הוא היא אני אתה מי איך כמה למה "
            "the a an of for with and or to in on is are what how do you know about my me".split())
_cache: dict = {}   # path -> (mtime, cols, rows)


def _nfc(s):
    return unicodedata.normalize("NFC", s) if isinstance(s, str) else ("" if s is None else str(s))


def _tokens(s: str) -> list[str]:
    s = _nfc(s)
    s = "".join(ch for ch in s if not (0x0591 <= ord(ch) <= 0x05C7))
    return [t for t in re.split(r"[^\wא-ת]+", s.lower()) if t and t not in _STOP and len(t) > 1]


# ── Canonical alias-normalization layer (SINGLE source of truth) ──────────────
# Explicit, curated aliases from OBSERVED live failures only. NOT stemming, NOT
# root extraction, NOT fuzzy matching — a literal phrase→canonical-phrase map.
# Applied before classify() AND before master token comparison so a spoken
# variant reaches the canonical term the machinery already understands.
_CONCEPT_ALIASES: dict[str, str] = {
    "דיסונאנס": "צרימה",       # transliteration → native music-theory term
    "פתיח יום": "פתיחת יום",   # colloquial short form → canonical day-opening phrase
}


def normalize_query(q: str) -> str:
    if not q:
        return q
    out = _nfc(q)
    for alias, canon in _CONCEPT_ALIASES.items():
        if alias in out:
            out = out.replace(alias, canon)
    return out


def human_master_path() -> Path | None:
    """Highest-version MASTER-PRODUCTION xlsx that has the canonical schema."""
    if not _HUMAN_DIR.exists():
        return None
    best = None
    for f in os.listdir(_HUMAN_DIR):
        n = _nfc(f)
        if "MASTER-PRODUCTION" not in n or not f.endswith(".xlsx") or f.startswith("~$"):
            continue
        m = re.search(r"MASTER-PRODUCTION-(\d+)\.(\d+)", n)
        if not m:
            continue
        ver = (int(m.group(1)), int(m.group(2)))
        if best is None or ver > best[0]:
            best = (ver, _HUMAN_DIR / f)
    return best[1] if best else None


def music_master_path() -> Path | None:
    return _MUSIC_FILE if _MUSIC_FILE.exists() else None


def _load(path: Path, sheet: str):
    """Load (cols, rows[list[dict]]) from a master sheet, cached by mtime.
    Returns (None, None) on any failure (evicted file, bad sheet, etc.)."""
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return None, None
    cached = _cache.get(str(path))
    if cached and cached[0] == mtime:
        return cached[1], cached[2]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close(); return None, None
        it = wb[sheet].iter_rows(values_only=True)
        cols = [_nfc(c) for c in next(it)]
        idx = {c: i for i, c in enumerate(cols)}
        rows = []
        for r in it:
            if not any(v is not None for v in r):
                continue
            rows.append({c: _nfc(r[i]) if i < len(r) else "" for c, i in idx.items()})
        wb.close()
    except Exception:
        return None, None
    _cache[str(path)] = (mtime, cols, rows)
    return cols, rows


def _query_master(path: Path | None, sheet: str, query: str, *, limit: int, kind: str) -> dict:
    """Return an UNKNOWN-aware retrieval dict for a master workbook."""
    if path is None:
        return {"status": "UNKNOWN", "reason": f"{kind} master file not found on disk",
                "path": None, "rows": [], "total": 0}
    cols, rows = _load(path, sheet)
    if cols is None:
        return {"status": "UNKNOWN", "reason": f"{kind} master unreadable (offline/evicted or bad sheet)",
                "path": str(path), "rows": [], "total": 0}
    missing = [c for c in _REQUIRED if c not in cols]
    if missing:
        return {"status": "UNKNOWN", "reason": f"{kind} master missing required columns {missing}",
                "path": str(path), "rows": [], "total": len(rows)}

    # only resolved, content-bearing units are answerable knowledge
    def _has_letter(s):
        return any(("א" <= ch <= "ת") or ("a" <= ch.lower() <= "z") for ch in _nfc(s))
    def resolved(r):
        st = r.get("Semantic_State", "")
        return (st in ("", "RESOLVED")) and _has_letter(r.get("Original_Text", ""))
    pool = [r for r in rows if resolved(r)]
    qtok = set(_tokens(normalize_query(query)))
    text_cols = [c for c in ("Original_Text", "Canonical_Text", "Heading", "Section", "Category") if c in cols]

    scored = []
    for r in pool:
        rtok = set()
        for c in text_cols:
            rtok |= set(_tokens(r.get(c, "")))
        score = len(qtok & rtok)
        if score:
            scored.append((score, r))
    scored.sort(key=lambda x: -x[0])
    # 2026-08-13: dedupe by Canonical_ID, keeping the higher-scored occurrence
    # of a cross-duplicate pair (the master's own Duplicate_Group tracks these
    # intentionally — the audit trail is correct; showing the same canonical
    # content twice in a 6-slot retrieval just wastes context budget). See
    # HUMAN-CONFIG-AUDIT-ADDENDUM-2026-08-13.md finding #2.
    seen_canonical: set[str] = set()
    deduped = []
    for _score, r in scored:
        cid = r.get("Canonical_ID", "")
        if cid and cid in seen_canonical:
            continue
        if cid:
            seen_canonical.add(cid)
        deduped.append(r)
    hits = deduped[:limit]
    representative = not hits
    if representative:                       # general "what do you know" query -> a small sample
        hits = pool[:limit]

    def prov(r):
        return {"canonical_id": r.get("Canonical_ID", ""), "atomic_id": r.get("Atomic_ID", ""),
                "source_id": r.get("Source_ID", ""), "heading": r.get("Heading", ""),
                "text": _nfc(r.get("Original_Text", ""))[:200]}
    return {"status": "LOADED", "path": str(path), "sheet": sheet,
            "total": len(pool), "representative": representative,
            "rows": [prov(r) for r in hits]}


def retrieve_human_master(query: str, *, limit: int = 6) -> dict:
    return _query_master(human_master_path(), _HUMAN_SHEET, query, limit=limit, kind="Human")


def retrieve_music_master(query: str, *, limit: int = 6) -> dict:
    return _query_master(music_master_path(), _MUSIC_SHEET, query, limit=limit, kind="Music")


def _best_overlap(path: Path | None, sheet: str, qtok: set) -> int:
    """Best token-overlap score of qtok against a master's RESOLVED units.
    0 when the master is unavailable or nothing overlaps. Same tokenization /
    resolved-unit rule as _query_master, so the probe and the real retrieval
    agree on what counts as a content match."""
    if path is None or not qtok:
        return 0
    cols, rows = _load(path, sheet)
    if cols is None or any(c not in cols for c in _REQUIRED):
        return 0

    def _has_letter(s):
        return any(("א" <= ch <= "ת") or ("a" <= ch.lower() <= "z") for ch in _nfc(s))
    text_cols = [c for c in ("Original_Text", "Canonical_Text", "Heading", "Section", "Category") if c in cols]
    best = 0
    for r in rows:
        st = r.get("Semantic_State", "")
        if not ((st in ("", "RESOLVED")) and _has_letter(r.get("Original_Text", ""))):
            continue
        rtok = set()
        for c in text_cols:
            rtok |= set(_tokens(r.get(c, "")))
        s = len(qtok & rtok)
        if s > best:
            best = s
    return best


def master_probe_scores(query: str) -> dict:
    """Zero-cue-match router fallback: how strongly a query's CONTENT tokens
    match each master's units. A distinctive word (e.g. 'דיסונאנס') scores > 0 in
    the master that contains it and 0 in the unrelated one — a domain signal
    without any keyword list. Returns {'music_config': int, 'human_config': int};
    {0,0} means no master signal (stay GENERAL)."""
    qtok = set(_tokens(normalize_query(query)))
    return {
        "music_config": _best_overlap(music_master_path(), _MUSIC_SHEET, qtok),
        "human_config": _best_overlap(human_master_path(), _HUMAN_SHEET, qtok),
    }
